"""
Import existing Excel data into Finance Desk database.
Run this once to import your 59 records from ALL_RECORDS.xlsx
"""
import os
import sys
import openpyxl
from datetime import datetime

# Add current directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import database as db

EXCEL_FILE = "Excel/ALL_RECORDS.xlsx"

def import_excel_to_db():
    """Import all records from Excel to SQLite database."""
    
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Error: Excel file not found at {EXCEL_FILE}")
        return False
    
    try:
        # Load Excel file
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        print(f"📋 Found headers: {headers}")
        
        # Read all data rows
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):  # Skip empty rows
                records.append(row)
        
        print(f"📊 Found {len(records)} records to import")
        
        # Import each record
        imported = 0
        skipped = 0
        
        for idx, row in enumerate(records, start=1):
            try:
                # Map Excel columns to database fields
                # Excel columns: SR NO, BID DATE, INVOICE NO, NAME, XCELL, PRODUCT, 
                # SERIAL NO / IMEI, PRICE, EMI, DI, BID, DP TAKEN, SCHEME, 
                # ACTUAL PRODUCT, GIVEN PROD PRICE, PHONE, ALT PHONE
                
                sr_no = str(row[0]) if row[0] else str(idx)
                bid_date = str(row[1]) if row[1] else ""
                invoice_no = str(row[2]) if row[2] else f"IMP-{idx:03d}"
                name = str(row[3]) if row[3] else ""
                xcell = str(row[4]) if row[4] else ""
                product = str(row[5]) if row[5] else ""
                serial_no = str(row[6]) if row[6] else ""
                price = float(row[7]) if row[7] else 0
                emi = float(row[8]) if row[8] else 0
                di = float(row[9]) if row[9] else 0
                bid = str(row[10]) if row[10] else ""
                dp_taken = float(row[11]) if row[11] else 0
                scheme = str(row[12]) if row[12] else ""
                actual_product = str(row[13]) if row[13] else ""
                given_prod_price = float(row[14]) if row[14] else 0
                phone = str(row[15]) if row[15] else ""
                alt_phone = str(row[16]) if len(row) > 16 and row[16] else ""
                
                # Prepare data for database
                data = {
                    "bid_date": bid_date,
                    "name": name,
                    "product": product,
                    "price": str(price),
                    "mobile": phone,
                    "address": "",
                    "bid": bid,
                    "emi": str(emi),
                    "di": str(di),
                    "scheme": scheme,
                }
                
                # Check if invoice already exists
                if invoice_no and db.check_invoice_exists(invoice_no):
                    print(f"⚠️  Skipping record {idx}: Invoice {invoice_no} already exists")
                    skipped += 1
                    continue
                
                # Check if serial already exists
                if serial_no and db.check_serial_exists(serial_no):
                    print(f"⚠️  Skipping record {idx}: Serial {serial_no} already exists")
                    skipped += 1
                    continue
                
                # Add to database
                record_id = db.add_record(
                    invoice_no=invoice_no,
                    data=data,
                    serial_no=serial_no,
                    xcell=xcell,
                    dp_taken=str(dp_taken),
                    product_given=actual_product,
                    given_prod_price=str(given_prod_price),
                    alt_phone=alt_phone,
                )
                
                imported += 1
                print(f"✅ Imported record {idx}/{len(records)}: {name} (Invoice: {invoice_no})")
                
            except Exception as e:
                print(f"❌ Error importing record {idx}: {str(e)}")
                skipped += 1
                continue
        
        print(f"\n🎉 Import Complete!")
        print(f"✅ Successfully imported: {imported} records")
        print(f"⚠️  Skipped: {skipped} records")
        print(f"\n📊 Total records in database: {len(db.load_all_records())}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error reading Excel file: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    print("=" * 60)
    print("📥 EXCEL TO DATABASE IMPORTER")
    print("=" * 60)
    print(f"Source: {EXCEL_FILE}")
    print(f"Target: Finance Desk Database")
    print("=" * 60)
    print()
    
    confirm = input("⚠️  This will import all records from Excel to database. Continue? (y/n): ")
    
    if confirm.lower() == 'y':
        print()
        success = import_excel_to_db()
        
        if success:
            print("\n✅ Import completed successfully!")
            print("💡 You can now run 'streamlit run app.py' to see your data")
        else:
            print("\n❌ Import failed. Please check the errors above.")
    else:
        print("❌ Import cancelled by user.")