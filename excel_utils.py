"""
Excel export utilities for Gagan's Finance Desk.
"""
import logging
import os
import re
from datetime import datetime
from typing import Dict, List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from config import EXCEL_FILE, HEADERS
from helpers import _normalize_date, _parse_date
from database import month_sort_key
import database as db


# Hoisted style objects - created once and reused across all cells/rows.
# Identical values to the originals; openpyxl serializes styles by value, so
# the saved workbook is byte-equivalent in appearance/behavior.
_LIGHT_ROW_FILL = PatternFill("solid", start_color="DCE6F1")
_WHITE_ROW_FILL = PatternFill("solid", start_color="FFFFFF")
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


def _sort_records_by_date_invoice(records):
    def sort_key(record):
        dt = _parse_date(record.get("bid_date", "")) or datetime.min
        inv = str(record.get("invoice_no") or "")
        match = re.search(r"(\d+)$", inv)
        inv_num = int(match.group(1)) if match else 0
        return (dt, inv_num)
    return sorted(records, key=sort_key)


def _apply_alternating_rows(ws, light_color="DCE6F1"):
    # Reuse the hoisted style objects (created once) instead of constructing a
    # new PatternFill per row + new Alignment per cell (the old hot path created
    # ~1.8M objects at 100K rows). Behavior/appearance are identical.
    light_fill = _LIGHT_ROW_FILL if light_color == "DCE6F1" else PatternFill("solid", start_color=light_color)
    white_fill = _WHITE_ROW_FILL
    center_align = _CENTER_ALIGN
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = light_fill if row_idx % 2 == 0 else white_fill
        for cell in row:
            cell.fill = fill
            cell.alignment = center_align


def style_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def update_excel_file():
    records = db.load_all_records()
    if records:
        wb = export_to_excel(records)
        wb.save(EXCEL_FILE)
        logging.info(f"Excel updated: {len(records)} records")


def export_to_excel(records):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    months_data: Dict[str, list] = {}
    for record in records:
        bid_date = str(record.get("bid_date", "") or "").strip()
        dt = _parse_date(bid_date)
        month = dt.strftime("%B_%Y").upper() if dt else (record.get("month", "") or "UNKNOWN")
        months_data.setdefault(month, []).append(record)
    sorted_months = sorted(months_data.keys(), key=lambda m: (-month_sort_key(m)[0], -month_sort_key(m)[1]))
    for month in sorted_months:
        month_records = months_data[month]
        month_records = _sort_records_by_date_invoice(month_records)
        ws = wb.create_sheet(title=month)
        ws.append(HEADERS)
        style_header(ws)
        for idx, record in enumerate(month_records, start=1):
            ws.append([
                idx, _normalize_date(record.get("bid_date", "")),
                record.get("invoice_no", ""), record.get("name", ""),
                record.get("xcell", ""), record.get("product", ""),
                record.get("serial_no", ""), record.get("price", ""),
                record.get("emi", ""), record.get("di", ""),
                record.get("bid", ""), record.get("dp_taken", ""),
                record.get("scheme", ""), record.get("actual_product", ""),
                record.get("given_prod_price", ""), record.get("phone", ""),
                record.get("alt_phone", ""), record.get("remarks", ""),
            ])
        _apply_alternating_rows(ws)
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_length + 2, 12), 45)
    return wb